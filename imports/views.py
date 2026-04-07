import csv
import io
from datetime import datetime

from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response

from victims.models import Victim
from victims.serializers import VictimWriteSerializer
from .models import BulkImport, BulkImportRow
from .serializers import BulkImportSerializer, BulkImportListSerializer

# Columns expected in the upload sheet (maps sheet header → Victim field)
COLUMN_MAP = {
    'full_name':        'full_name',
    'gender':           'gender',
    'age_at_death':     'age_at_death',
    'community_ward':   'community_ward',
    'year_of_death':    'year_of_death',
    'date_of_death':    'date_of_death',
    'cause_of_death':   'cause_of_death',
    'biographical_note':'biographical_note',
    'home_lat':         'home_lat',
    'home_lng':         'home_lng',
    'incident_lat':     'incident_lat',
    'incident_lng':     'incident_lng',
    'burial_lat':       'burial_lat',
    'burial_lng':       'burial_lng',
    'source':           'source',
    'consent_status':   'consent_status',
    'consent_notes':    'consent_notes',
}

REQUIRED_COLUMNS = {'full_name', 'community_ward', 'source'}


def parse_file(file):
    """Return list of row dicts from CSV or Excel upload."""
    name = file.name.lower()
    if name.endswith('.csv'):
        decoded = file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        return [dict(row) for row in reader]
    elif name.endswith(('.xlsx', '.xls')):
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue  # skip empty rows
            rows.append({headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)})
        return rows
    else:
        raise ValueError('Unsupported file type. Please upload a .csv or .xlsx file.')


def clean_row(raw):
    """Normalise raw string values to Python types for the serializer."""
    nullify = lambda v: None if (v is None or str(v).strip() == '') else str(v).strip()
    data = {}
    for sheet_col, model_field in COLUMN_MAP.items():
        val = nullify(raw.get(sheet_col))
        data[model_field] = val

    # Coerce numeric fields
    for f in ('age_at_death', 'year_of_death', 'home_lat', 'home_lng',
              'incident_lat', 'incident_lng', 'burial_lat', 'burial_lng'):
        if data.get(f) is not None:
            try:
                data[f] = float(data[f]) if '.' in str(data[f]) else int(data[f])
            except (ValueError, TypeError):
                data[f] = data[f]  # leave as-is; serializer will catch the error

    # Default consent status
    if not data.get('consent_status'):
        data['consent_status'] = 'PENDING'

    return data


def validate_row(raw, request):
    cleaned = clean_row(raw)
    s = VictimWriteSerializer(data=cleaned, context={'request': request})
    if s.is_valid():
        return True, {}, cleaned
    return False, {k: [str(e) for e in v] for k, v in s.errors.items()}, cleaned


class BulkImportViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAdminUser]
    http_method_names  = ['get', 'post', 'head', 'options']

    def get_queryset(self):
        if self.request.user.is_superuser:
            return BulkImport.objects.all()
        return BulkImport.objects.filter(uploaded_by=self.request.user)

    def get_serializer_class(self):
        if self.action == 'list':
            return BulkImportListSerializer
        return BulkImportSerializer

    def create(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            raw_rows = parse_file(file)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if not raw_rows:
            return Response({'error': 'The file is empty.'}, status=status.HTTP_400_BAD_REQUEST)

        # Check required columns are present
        headers = set(raw_rows[0].keys())
        missing = REQUIRED_COLUMNS - headers
        if missing:
            return Response(
                {'error': f'Missing required columns: {", ".join(sorted(missing))}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        bulk_import = BulkImport.objects.create(
            file=file,
            original_filename=file.name,
            uploaded_by=request.user,
            total_rows=len(raw_rows),
        )

        # Validate all rows and store them
        valid_count = 0
        row_objects = []
        for i, raw in enumerate(raw_rows, start=2):
            is_valid, errors, _ = validate_row(raw, request)
            if is_valid:
                valid_count += 1
            row_objects.append(BulkImportRow(
                batch=bulk_import,
                row_number=i,
                raw_data=raw,
                is_valid=is_valid,
                validation_errors=errors,
            ))

        BulkImportRow.objects.bulk_create(row_objects)
        bulk_import.valid_rows = valid_count
        bulk_import.save()

        serializer = BulkImportSerializer(bulk_import, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        if not request.user.is_superuser:
            return Response({'error': 'Only superadmins may approve imports.'}, status=status.HTTP_403_FORBIDDEN)

        bulk_import = self.get_object()
        if bulk_import.status != 'PENDING':
            return Response({'error': f'Import is already {bulk_import.status}.'}, status=status.HTTP_400_BAD_REQUEST)

        notes    = request.data.get('review_notes', '')
        created  = 0
        skipped  = 0

        for row in bulk_import.rows.filter(is_valid=True, victim__isnull=True):
            _, _, cleaned = validate_row(row.raw_data, request)
            cleaned['added_by'] = request.user
            # Force to PENDING consent unless explicitly set in sheet
            if cleaned.get('consent_status') not in ('CONSENTED', 'ANONYMOUS', 'PENDING'):
                cleaned['consent_status'] = 'PENDING'
            try:
                victim = Victim.objects.create(**cleaned)
                row.victim = victim
                row.save()
                created += 1
            except Exception:
                skipped += 1

        bulk_import.status      = 'APPROVED'
        bulk_import.reviewed_by = request.user
        bulk_import.reviewed_at = timezone.now()
        bulk_import.review_notes = notes
        bulk_import.save()

        return Response({'created': created, 'skipped': skipped})

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        if not request.user.is_superuser:
            return Response({'error': 'Only superadmins may reject imports.'}, status=status.HTTP_403_FORBIDDEN)

        bulk_import = self.get_object()
        if bulk_import.status != 'PENDING':
            return Response({'error': f'Import is already {bulk_import.status}.'}, status=status.HTTP_400_BAD_REQUEST)

        bulk_import.status       = 'REJECTED'
        bulk_import.reviewed_by  = request.user
        bulk_import.reviewed_at  = timezone.now()
        bulk_import.review_notes = request.data.get('review_notes', '')
        bulk_import.save()

        return Response({'status': 'rejected'})

    @action(detail=False, methods=['get'], url_path='template')
    def template(self, request):
        """Return a downloadable CSV template with the correct column headers."""
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="mapping_memory_import_template.csv"'
        writer = csv.writer(response)
        writer.writerow(list(COLUMN_MAP.keys()))
        writer.writerow([
            'Emmanuel Yakubu', 'M', '34', 'Angwan Rogo', '2001', '',
            'Shot during violence', 'Brief life story here.',
            '9.917', '8.896', '', '', '', '', 'Family interview 2024',
            'CONSENTED', '',
        ])
        return response
